from openpyxl import Workbook

def export_day_report(report, output_path, date_str):
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Data", date_str])
    ws_summary.append(["Total de Vendas", report["summary"]["total_sales"]])
    ws_summary.append(["Total Pago", report["summary"]["total_paid"]])
    ws_summary.append(["Total Pendente", report["summary"]["total_pending"]])

    ws_sales = wb.create_sheet("Vendas")
    ws_sales.append(["Cliente", "Venda", "Data", "Total", "Modalidade", "Parcelas", "Pago", "Valor Pago", "Pendente"])
    for row in report["sales_rows"]:
        payment_method = row.get("payment_method", "N/A")
        installments = row.get("installments", 1)
        installments_text = f"{installments}x" if installments > 1 else "-"
        ws_sales.append([
            row["client"],
            row["sale_id"],
            row["date"],
            row["total"],
            payment_method,
            installments_text,
            "Sim" if row["paid"] else "Não",
            row["paid_amount"],
            row["remaining"]
        ])

    ws_products = wb.create_sheet("Produtos")
    ws_products.append(["Produto", "Quantidade", "Total"])
    for row in report["products_rows"]:
        ws_products.append([row["name"], row["qty"], row["total"]])

    wb.save(output_path)
